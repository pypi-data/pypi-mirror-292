import re
from pathlib import Path

from openpyxl import load_workbook


def _dictionarise(column_names, row):
    return dict(
        (name, str(cell.value))
        for name, cell in zip(column_names, row)
        if name and cell.value)


def _normalise_cell(cell):
    if isinstance(cell, str):
        return re.sub(r'\s+', ' ', cell.strip())
    return cell


def normalise_whitespace(row):
    """Return table row with normalised white space.

    This involves stripping leading and trailing whitespace, as well as
    consolidating white space to single spaces.
    """
    pairs = (
        (k, _normalise_cell(v))
        for k, v in row.items())
    return {
        k: v for k, v in pairs
        if not isinstance(v, str) or v}


def sheet_to_list(worksheet):
    """Return an Excel worksheet as list of dictionaries.

    Each dictionary in the list represents a row of the worksheet.  The
    dictionaries themselvs map column names to values.
    """
    column_names = [cell.value for cell in worksheet['1']]
    data = worksheet['2:{}'.format(worksheet.max_row)]
    if worksheet.max_row == 2:
        # if you pass the range '2:2' into openpyxl it only returns a flat tuple
        # of cells.
        rows = [_dictionarise(column_names, data)]
    else:
        rows = [_dictionarise(column_names, row) for row in data]
    return [row for row in rows if row]


def load_excel_data(folder):  # pragma: nocover
    """Return all Excel workbooks in `folder`.

    This returns a dictionary mapping the basenames of the Excel files to
    their data.

    Note: This assumes that every workbook consists of exactly one work sheet
    and that the first row of this sheet contains the column names.
    """
    data = {}
    for filename in Path(folder).glob('*.xlsx'):
        workbook = load_workbook(filename=str(filename))
        if not workbook.sheetnames:
            continue
        table_name = filename.stem
        data[table_name] = sheet_to_list(workbook[workbook.sheetnames[0]])
    data = {
        table_name: list(map(normalise_whitespace, table))
        for table_name, table in data.items()}
    return data
