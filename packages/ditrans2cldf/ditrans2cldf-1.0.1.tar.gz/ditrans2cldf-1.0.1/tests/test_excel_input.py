import unittest
from ditrans2cldf import excel_input
from openpyxl import Workbook


class WorkbookConversion(unittest.TestCase):

    def test_all_cells_filled(self):
        workbook = Workbook()
        sheet = workbook.create_sheet()
        sheet.append(['col1', 'col2', 'col3'])
        sheet.append(['val1.1', 'val1.2', 'val1.3'])
        sheet.append(['val2.1', 'val2.2', 'val2.3'])
        result = excel_input.sheet_to_list(sheet)
        expected = [
            {'col1': 'val1.1', 'col2': 'val1.2', 'col3': 'val1.3'},
            {'col1': 'val2.1', 'col2': 'val2.2', 'col3': 'val2.3'},
        ]
        self.assertEqual(result, expected)

    def test_ignore_empty_cells(self):
        workbook = Workbook()
        sheet = workbook.create_sheet()
        sheet.append(['col1', 'col2', 'col3'])
        sheet.append(['val1.1', '', 'val1.3'])
        sheet.append(['val2.1', 'val2.2', ''])
        result = excel_input.sheet_to_list(sheet)
        expected = [
            {'col1': 'val1.1', 'col3': 'val1.3'},
            {'col1': 'val2.1', 'col2': 'val2.2'},
        ]
        self.assertEqual(result, expected)

    def test_ignore_empty_rows(self):
        workbook = Workbook()
        sheet = workbook.create_sheet()
        sheet.append(['col1', 'col2', 'col3'])
        sheet.append(['', '', ''])
        sheet.append(['val2.1', 'val2.2', 'val2.3'])
        result = excel_input.sheet_to_list(sheet)
        expected = [
            {'col1': 'val2.1', 'col2': 'val2.2', 'col3': 'val2.3'},
        ]
        self.assertEqual(result, expected)

    def test_ignore_nameless_columns(self):
        workbook = Workbook()
        sheet = workbook.create_sheet()
        sheet.append(['col1', '', 'col3'])
        sheet.append(['val1.1', 'val1.2', 'val1.3'])
        sheet.append(['val2.1', 'val2.2', 'val2.3'])
        result = excel_input.sheet_to_list(sheet)
        expected = [
            {'col1': 'val1.1', 'col3': 'val1.3'},
            {'col1': 'val2.1', 'col3': 'val2.3'},
        ]
        self.assertEqual(result, expected)

    def test_only_one_row_in_the_table(self):
        workbook = Workbook()
        sheet = workbook.create_sheet()
        sheet.append(['col1', 'col2', 'col3'])
        sheet.append(['val1.1', 'val1.2', 'val1.3'])
        result = excel_input.sheet_to_list(sheet)
        expected = [
            {'col1': 'val1.1', 'col2': 'val1.2', 'col3': 'val1.3'},
        ]
        self.assertEqual(result, expected)


class CellNormalisation(unittest.TestCase):

    def test_consolidate_whitespace(self):
        row = {'col1': 'val  1', 'col2': 'val \t \n 2'}
        result = excel_input.normalise_whitespace(row)
        expected = {'col1': 'val 1', 'col2': 'val 2'}
        self.assertEqual(result, expected)

    def test_strip_whitespace(self):
        row = {'col1': ' val 1 ', 'col2': '\tval 2\n'}
        result = excel_input.normalise_whitespace(row)
        expected = {'col1': 'val 1', 'col2': 'val 2'}
        self.assertEqual(result, expected)

    def test_strip_empty_cells(self):
        row = {'col1': 'val 1', 'col2': '   '}
        result = excel_input.normalise_whitespace(row)
        expected = {'col1': 'val 1'}
        self.assertEqual(result, expected)

    def test_only_affect_strings(self):
        row = {'col1': 0, 'col2': 1}
        result = excel_input.normalise_whitespace(row)
        expected = {'col1': 0, 'col2': 1}
        self.assertEqual(result, expected)
