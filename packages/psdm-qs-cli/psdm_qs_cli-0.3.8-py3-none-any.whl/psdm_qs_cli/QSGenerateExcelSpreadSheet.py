#!/usr/bin/env python
"""
Use the questionnaire client to generate a Excel spreadsheet.
We take a list of attributes in a file, (names with mappings) and generate a
row per proposal and a column per attribute.
"""

import argparse
import json
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, colors

from psdm_qs_cli import QuestionnaireClient

logging.basicConfig(level=logging.DEBUG)

logger = logging.getLogger(__name__)


def generateExcelSpreadSheetForRun(qs, run, attributes_file, excelFilePath):
    """
    Generate a Excel spreadsheet with data from a run.
    :param: qs - A Questionnaire client
    :run: The number number/name; this is a string like run15 which is what the
        questionnaire uses in its URL
    """
    column2Names = [("proposal_id", "Proposal")]
    with open(attributes_file) as f:
        attrs = json.load(f)
        column2Names.extend((x["attr"], x["label"]) for x in attrs)

    wb = Workbook()
    ws = wb.active
    ws.title = run
    fontStyle = Font(name="Times New Roman", size=12, color=colors.BLACK)

    # Generate the header column using the 2 element in the column2Names tuples
    for clnum in range(len(column2Names)):
        cl = ws.cell(row=1, column=clnum + 1, value=column2Names[clnum][1])
        cl.font = fontStyle

    # Get a list of proposals
    proposals = qs.getProposalsListForRun(run)
    rowNum = 1
    for proposalid in sorted(proposals.keys()):
        print("Getting details for proposal ", proposalid)
        proposalDetails = qs.getProposalDetailsForRun(run, proposalid)
        # Add the details of each proposal to the information obtained from the proposal list call.
        proposals[proposalid].update(proposalDetails)
        for clnum in range(len(column2Names)):
            ckey = column2Names[clnum][0]
            if ckey in proposals[proposalid]:
                _ = ws.cell(
                    row=rowNum + 1, column=clnum + 1, value=proposals[proposalid][ckey]
                )
            else:
                _ = ws.cell(row=rowNum + 1, column=clnum + 1, value="")
        rowNum = rowNum + 1

    wb.save(excelFilePath)
    print("Saved data into", excelFilePath)


def main():
    parser = argparse.ArgumentParser(
        description="Load data from the questionnaire into a an Excel spreadsheet"
    )
    parser.add_argument(
        "--questionnaire_url",
        default="https://pswww.slac.stanford.edu/ws-kerb/questionnaire",
    )
    parser.add_argument("--no_kerberos", action="store_false")
    parser.add_argument("--user")
    parser.add_argument("--password")
    parser.add_argument("run")
    parser.add_argument(
        "attributes_file",
        help="A JSON file with an array of dicts; each of which has a attrname and a label.",
    )
    parser.add_argument("excelFilePath")
    args = parser.parse_args()

    qs = QuestionnaireClient(
        args.questionnaire_url, args.no_kerberos, user=args.user, pw=args.password
    )
    generateExcelSpreadSheetForRun(
        qs, args.run, args.attributes_file, args.excelFilePath
    )


if __name__ == "__main__":
    main()
